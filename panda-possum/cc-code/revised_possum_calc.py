# -*- coding: utf-8 -*-
"""
P.O.S.S.U.M. (PANDA Optimization Supplemental Script Utilizing Machine-learnig) 

This script is designed to make it easier to iterate and change canoe parameters for PANDA (Program for Automated Naval Design Analysis).
This is done by using the bisection method on the inputsetup.txt file required for bulk canoe analysis in PANDA.

Created in 2019 by Rick Liu for University of Toronto Concrete Canoe Team

Revised in 2025-2026 by Lucas and Elorie as part of PANDA overhaul
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent / "data"

class PossumCalc:
    
    def new_length(self, canoe_array, max_canoe):
        length = canoe_array[0]
        lower_limit=float(length[1])  
        upper_limit=float(length[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.05:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[0])
        iterations=2
        return lower_limit, upper_limit, iterations
    
    def new_lp(self, canoe_array, max_canoe):
        lp=canoe_array[1]
        lower_limit=float(lp[1])  
        upper_limit=float(lp[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.05:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[1])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_ld(self, canoe_array, max_canoe):
        ld=canoe_array[2]
        lower_limit=float(ld[1])  
        upper_limit=float(ld[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.05:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[2])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_lf(self, canoe_array, max_canoe):
        lf=canoe_array[3]
        lower_limit=float(lf[1])  
        upper_limit=float(lf[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.05:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[3])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_width(self, canoe_array, max_canoe):
        width=canoe_array[4]
        lower_limit=float(width[1])  
        upper_limit=float(width[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.01:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[4])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_t1(self, canoe_array, max_canoe):
        t1=canoe_array[5]
        lower_limit=float(t1[1])  
        upper_limit=float(t1[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.01:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[5])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_t2(self, canoe_array, max_canoe):
        t2=canoe_array[6]
        lower_limit=float(t2[1])  
        upper_limit=float(t2[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.01:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[6])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_depth(self, canoe_array, max_canoe):
        depth=canoe_array[7]
        lower_limit=float(depth[1])  
        upper_limit=float(depth[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.01:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[7])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_bowrocker(self, canoe_array, max_canoe):
        bowrocker=canoe_array[8]
        lower_limit=float(bowrocker[1])  
        upper_limit=float(bowrocker[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.01:
            iterations=1
            return lower_limit, lower_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[8])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_sternrocker(self, canoe_array, max_canoe):
        sternrocker=canoe_array[9]
        lower_limit=float(sternrocker[1])  
        upper_limit=float(sternrocker[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.01:
            iterations=1
            return lower_limit, lower_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[9])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_flareangle(self, canoe_array, max_canoe):
        flareangle=canoe_array[10]
        lower_limit=float(flareangle[1])  
        upper_limit=float(flareangle[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.02:
            iterations=1
            return lower_limit, lower_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[10])
        iterations=2
        return lower_limit, upper_limit, iterations

    def new_shapeparam(self, canoe_array, max_canoe):        # Shape parameter 'n'
        shapeparam=canoe_array[11]
        lower_limit=float(shapeparam[1])  
        upper_limit=float(shapeparam[2])  
        delta=abs(upper_limit-lower_limit)
        if delta<0.05:
            iterations=1
            return upper_limit, upper_limit, iterations
        lower_limit, upper_limit=self.bisection(lower_limit, upper_limit, max_canoe[11])
        iterations=2
        return lower_limit, upper_limit, iterations
        
    def bisection(self, lower_limit, upper_limit, canoe):    # Actual search function, also known as golden section search
        lower_delta=abs(canoe-lower_limit)
        upper_delta=abs(canoe-upper_limit)
        delta=upper_limit-lower_limit
        if lower_delta>upper_delta:
            new_upper=upper_limit
            new_lower=lower_limit+(delta-(delta/1.618))       # 1.618 is the golden ratio
        else:
            new_lower=lower_limit
            new_upper=upper_limit-(delta-(delta/1.618))
        return new_lower, new_upper
    
def main():        
        
    # process the input file                                  # Would need to change this to read and edit the xlsx
    input_setup = pd.read_excel(DATA_DIR / "input.xlsx", "canoe")

    # change input_setup to canoe_array
    canoe_array = []
    for index, row in input_setup.iterrows():
        canoe_array.append([row['Parameter'], row['Min'], row['Max'], row['Iterations']])
        
    # with open(Path("cc-code/data/inputtable.txt"), "r") as table:         # We can keep this, it appears to have some function, and wouldn't "replace" the xlsx
        # reader=csv.reader(table, delimiter = '\t')
        # inputtable=list(reader)
            
    # Read output.csv using pandas
    output_df = pd.read_csv(DATA_DIR / "output.csv")
    
    # Filter out failed analyses and find the row with max Average Score
    successful_df = output_df[output_df['Success'] == True].copy()
    
    if successful_df.empty:
        print("No successful canoe analyses in this iteration. Change input parameters in input.xlsx and try again.")
        return
    
    # Find the best canoe (highest Average Score)
    best_row = successful_df.loc[successful_df['Average Score'].idxmax()]
    max_score = best_row['Average Score']
    
    # Extract parameters from best canoe, convert into list of floats
    variant_str = best_row['Canoe Variant']
    variant_values = variant_str.strip('()').split(', ')
    
    max_canoe = []
    for v in variant_values:
        # Remove numpy wrapper text if present before converting to float
        clean_v = v.replace("np.float64(", "").replace(")", "")
        max_canoe.append(float(clean_v))
    
    print(f"Best canoe found with Average Score: {max_score}")
    print(f"Parameters: {max_canoe}")
           
        
    # num_rows = ['13']                     # I don't think this has any significance... doesn't appear to play into any other functions. In the original it only got written into the file
    possum = PossumCalc()                   # Create instance of PossumCalc. Without this, python expects a third input for possumcalc functions
    length=['Length', possum.new_length(canoe_array, max_canoe)[0], possum.new_length(canoe_array, max_canoe)[1], possum.new_length(canoe_array, max_canoe)[2]]
    lp=['Lp', possum.new_lp(canoe_array, max_canoe)[0], possum.new_lp(canoe_array, max_canoe)[1], possum.new_lp(canoe_array, max_canoe)[2]]
    ld=['Ld', possum.new_ld(canoe_array, max_canoe)[0], possum.new_ld(canoe_array, max_canoe)[1], possum.new_ld(canoe_array, max_canoe)[2]]
    lf=['Lf', possum.new_lf(canoe_array, max_canoe)[0], possum.new_lf(canoe_array, max_canoe)[1], possum.new_lf(canoe_array, max_canoe)[2]]
    width=['W', possum.new_width(canoe_array, max_canoe)[0], possum.new_width(canoe_array, max_canoe)[1], possum.new_width(canoe_array, max_canoe)[2]]
    t1=['t1', possum.new_t1(canoe_array, max_canoe)[0], possum.new_t1(canoe_array, max_canoe)[1], possum.new_t1(canoe_array, max_canoe)[2]]
    t2=['t2', possum.new_t2(canoe_array, max_canoe)[0], possum.new_t2(canoe_array, max_canoe)[1], possum.new_t2(canoe_array, max_canoe)[2]]
    depth=['d', possum.new_depth(canoe_array, max_canoe)[0], possum.new_depth(canoe_array, max_canoe)[1], possum.new_depth(canoe_array, max_canoe)[2]]
    bowrocker=['b', possum.new_bowrocker(canoe_array, max_canoe)[0], possum.new_bowrocker(canoe_array, max_canoe)[1], possum.new_bowrocker(canoe_array, max_canoe)[2]]
    sternrocker=['s', possum.new_sternrocker(canoe_array, max_canoe)[0], possum.new_sternrocker(canoe_array, max_canoe)[1], possum.new_sternrocker(canoe_array, max_canoe)[2]]
    flareangle=['f', possum.new_flareangle(canoe_array, max_canoe)[0], possum.new_flareangle(canoe_array, max_canoe)[1], possum.new_flareangle(canoe_array, max_canoe)[2]]
    shapeparam=['n', possum.new_shapeparam(canoe_array, max_canoe)[0], possum.new_shapeparam(canoe_array, max_canoe)[1], possum.new_shapeparam(canoe_array, max_canoe)[2]]
    density=['density', canoe_array[12][1], canoe_array[12][2], 1] # Density is not iterated upon, since it's a property decided by concrete team.

    new_canoe_array = []
    new_canoe_array.append(length)
    new_canoe_array.append(lp)
    new_canoe_array.append(ld)
    new_canoe_array.append(lf)
    new_canoe_array.append(width)
    new_canoe_array.append(t1)
    new_canoe_array.append(t2)
    new_canoe_array.append(depth)
    new_canoe_array.append(bowrocker)
    new_canoe_array.append(sternrocker)
    new_canoe_array.append(flareangle)
    new_canoe_array.append(shapeparam)
    new_canoe_array.append(density)

    print(new_canoe_array)

    # Old code for writing back into text file
    # inputsetup_out = [x for x in new_canoe_array if x]                  # Writing corrected min/max values and iterations back into the array
    # with open(Path('data/inputsetup.txt'), "w") as setup_out:
    #     writer=csv.writer(setup_out, delimiter = '\t')
    #     for item in inputsetup_out:
    #         writer.writerow(item)
        
    output_df = pd.DataFrame(new_canoe_array, columns=['Parameter', 'Min', 'Max', 'Iterations'])
    
    # Overwrite only the canoe sheet, leaving loadcase untouched
    input_path = DATA_DIR / "input.xlsx"
    input_excel = load_workbook(input_path)
    
    # Remove existing canoe sheet
    if 'canoe' in input_excel.sheetnames:
        del input_excel['canoe']
    
    # Create new canoe sheet and write data
    canoe_sheet = input_excel.create_sheet('canoe')
    
    # Write header row
    for col_idx, col_name in enumerate(output_df.columns, start=1):
        canoe_sheet.cell(row=1, column=col_idx, value=col_name)
    
    # Write data rows
    for row_idx, row in enumerate(output_df.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            canoe_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    input_excel.save(input_path)

    print("New canoe parameters for next iteration successfully written into input.xlsx\n")
    return

# Leftover from when POSSUM was its own standalone script
    # if __name__ == "__main__":
    #   import sys
    #   possum=PossumCalc()
    #   possum.main()
    #   sys.exit()
            